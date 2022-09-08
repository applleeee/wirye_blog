from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# 키워드 기존 통계파일에서 불러오기
wb = load_workbook("위례블로그 통계(국어,수학).xlsx")
ws1 = wb["국어키워드"]
ws2 = wb["수학키워드"]

# 키워드(국어, 수학 따로)
ko_keywords = [] 
math_keywords = []


for word in ws1.iter_rows(min_row=2, min_col=5):
    if word[0].value is not None:
        ko_keyword = word[0].value
        ko_keywords.append(ko_keyword)

for word in ws2.iter_rows(min_row=2, min_col=5):
    if word[0].value is not None:
        math_keyword = word[0].value
        math_keywords.append(math_keyword)


url = "https://search.naver.com/search.naver?where=view&sm=tab_jum&query="

browser = webdriver.Chrome("C:/Users/user/Desktop/코딩/wirye_blog/chromedriver.exe")
browser.get(url)


# 검색 및 결과값 입력
def input_result(keyword, sheet):
    rank = []
    title = []
    for keyword1 in keyword:
        elem = browser.find_element_by_id("nx_query")
        elem.clear()
        elem.send_keys(keyword1)
        elem = browser.find_element_by_class_name("bt_search")
        elem.click()
        soup = BeautifulSoup(browser.page_source, "lxml")

        blog = soup.find("a", {"href": "https://blog.naver.com/wredusky"})
        if blog:
            print("검색중")
            li = blog.find_parent("li")
            rank1 = blog.find_parent("li")["data-cr-rank"]
            title1 = li.find("a", {"class" : "api_txt_lines total_tit _cross_trigger"}).text
            rank.append(rank1)
            title.append(title1)
            
        else:
            print("순위없음")
            
            rank.append("순위없음")
            title.append("게시물 없음")

    # 오늘날짜 입력
    d = datetime.now()
    only_date = d.date()
    sheet["g1"] = str(only_date) 

    # 순위, 제목 입력
    for index, cell in enumerate(sheet.iter_rows(min_col=5, min_row=2)):
        
        cell[2].value = rank[index]
        cell[5].value = title[index]

    wb.save("위례블로그 통계(국어,수학).xlsx")


input_result(ko_keywords, ws1)
input_result(math_keywords, ws2)



# 1페이지 이상 키워드 빨간색 표시
def color_change(sheet):
    for a in sheet.iter_rows(min_col=5,min_row=2):
        if str(a[2].value) == "순위없음" or int(a[2].value) > 5:
            for red in range(0,6):
                a[red].fill = PatternFill(fgColor="FF6D6D", fill_type="solid") # 빨강

        elif int(a[2].value) <= 5:
            for blue in range(0,6):
                a[blue].fill = PatternFill(fgColor="D9E1F2", fill_type="solid") # 파랑


color_change(ws1)
color_change(ws2)




wb.save("위례블로그 통계(국어,수학).xlsx")
